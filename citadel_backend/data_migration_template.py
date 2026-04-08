import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import json
import os
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.colors import Color

# Hardcoded bank data instead of loading from bank.json
bank_names = [
    "Maybank Berhad",
    "Maybank Islamic Berhad",
    "CIMB Bank Berhad",
    "CIMB Islamic Bank Berhad",
    "Bank Islam Malaysia Berhad",
    "Bank Kerjasama Rakyat Malaysia Berhad",
    "RHB Bank Berhad",
    "RHB Islamic Bank Berhad",
    "Hong Leong Bank Berhad",
    "Hong Leong Islamic Bank Berhad",
    "Public Bank Berhad",
    "Public Islamic Bank Berhad",
    "Bank Simpanan Nasional",
    "HSBC Bank Malaysia Berhad",
    "Ambank (M) Berhad",
    "AmBank Islamic Berhad",
    "Al Rajhi Banking & Investment Corporation (M) Berhad",
    "Bank Muamalat Malaysia Berhad",
    "Standard Chartered Bank (M) Berhad",
    "Affin Bank Berhad",
    "Affin Islamic Bank Berhad",
    "United Overseas Bank Berhad",
    "OCBC Al-Amin Bank Berhad",
    "Alliance Bank Malaysia Berhad",
    "Alliance Islamic Bank Berhad",
    "OCBC Bank Berhad",
    "MBSB Bank Berhad",
    "Agro Bank",
    "GX Bank Berhad",
    "Bank of China (Malaysia) Berhad"
]

# Hardcoded swift codes for each bank
swift_codes = {
    "Maybank Berhad": "MBBEMYKL",
    "Maybank Islamic Berhad": "MBISMYKL",
    "CIMB Bank Berhad": "CIBBMYKL",
    "CIMB Islamic Bank Berhad": "CTBBMYKL",
    "Bank Islam Malaysia Berhad": "BIMBMYKL",
    "Bank Kerjasama Rakyat Malaysia Berhad": "BKRMMYKL",
    "RHB Bank Berhad": "RHBBMYKL",
    "RHB Islamic Bank Berhad": "RHBAMYKL",
    "Hong Leong Bank Berhad": "HLBBMYKL",
    "Hong Leong Islamic Bank Berhad": "HLIBMYKL",
    "Public Bank Berhad": "PBBEMYKL",
    "Public Islamic Bank Berhad": "PIBEMYK1",
    "Bank Simpanan Nasional": "BSNAMYK1",
    "HSBC Bank Malaysia Berhad": "HBMBMYKL",
    "Ambank (M) Berhad": "ARBKMYKL",
    "AmBank Islamic Berhad": "AISLMYKL",
    "Al Rajhi Banking & Investment Corporation (M) Berhad": "RJHIMYKL",
    "Bank Muamalat Malaysia Berhad": "BMMBMYKL",
    "Standard Chartered Bank (M) Berhad": "SCBLMYKX",
    "Affin Bank Berhad": "PHBMMYKL",
    "Affin Islamic Bank Berhad": "AIBBMYKL",
    "United Overseas Bank Berhad": "UOVBMYKL",
    "OCBC Al-Amin Bank Berhad": "OABBMYKL",
    "Alliance Bank Malaysia Berhad": "MFBBMYKL",
    "Alliance Islamic Bank Berhad": "ALSRMYKL",
    "OCBC Bank Berhad": "OCBCMYKL",
    "MBSB Bank Berhad": "AFBQMYKL",
    "Agro Bank": "AGOBMYKL",
    "GX Bank Berhad": "GXSPMYKL",
    "Bank of China (Malaysia) Berhad": "BKCHMYKL"
}

# Function to create a comma-separated string for data validation
def create_validation_string(items):
    return '"' + ','.join(items) + '"'

# Create a workbook with multiple sheets for different entity types
wb = Workbook()

# Get the active sheet (the default sheet)
default_sheet = wb.active

# Define some styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
mandatory_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
optional_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
section_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Helper function to create and style headers
def create_headers(ws, headers, descriptions, mandatory_fields):
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add descriptions in row 2
        desc_cell = ws.cell(row=2, column=col, value=descriptions[col-1])
        desc_cell.alignment = Alignment(wrap_text=True)
        
        # Mark mandatory fields
        if header in mandatory_fields:
            ws.cell(row=3, column=col, value="Mandatory")
            ws.cell(row=3, column=col).fill = mandatory_fill
        else:
            ws.cell(row=3, column=col, value="Optional")
            ws.cell(row=3, column=col).fill = optional_fill
    
    # Freeze panes to keep headers visible
    ws.freeze_panes = 'A4'
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

# Helper function to add sheet-specific data validation
def add_sheet_specific_validation(worksheet, validation_type, formula1, range_or_ranges, field_name=None, allow_blank=True):
    """
    Adds a data validation to a specific worksheet and cell range by creating a new DataValidation object.
    This prevents validation crossover between different sheets.
    
    Args:
        worksheet: The worksheet to add validation to
        validation_type: The type of validation (e.g., "list", "textLength")
        formula1: The formula or value for the validation
        range_or_ranges: A string or list of strings representing cell ranges (e.g., 'A4:A1000')
        field_name: Optional name of the field for documentation
        allow_blank: Whether to allow blank values (default True)
    """
    # Create a new DataValidation object specific to this worksheet
    new_validation = DataValidation(type=validation_type, formula1=formula1, allow_blank=allow_blank)
    
    # Add the validation to the worksheet
    worksheet.add_data_validation(new_validation)
    
    # Apply to single range or multiple ranges
    if isinstance(range_or_ranges, list):
        for cell_range in range_or_ranges:
            new_validation.add(cell_range)
    else:
        new_validation.add(range_or_ranges)

# Helper function to add data validation to a specific sheet and range
def add_validation_to_sheet(worksheet, validation, range_or_ranges, field_name=None):
    """
    Adds a data validation to a specific worksheet and cell range.
    
    Args:
        worksheet: The worksheet to add validation to
        validation: The DataValidation object
        range_or_ranges: A string or list of strings representing cell ranges (e.g., 'A4:A1000')
        field_name: Optional name of the field for documentation
    """
    # Add the validation to the worksheet
    worksheet.add_data_validation(validation)
    
    # Apply to single range or multiple ranges
    if isinstance(range_or_ranges, list):
        for cell_range in range_or_ranges:
            validation.add(cell_range)
    else:
        validation.add(range_or_ranges)

# ====================== INSTRUCTIONS SHEET ======================
ws_instructions = wb.create_sheet(title="Instructions", index=0)

instructions = [
    ["Data Migration Template - Instructions", ""],
    ["", ""],
    ["This Excel workbook contains templates for data migration of various entities and transactions in the Citadel system.", ""],
    ["", ""],
    ["Sheet Structure:", ""],
    ["", ""],
    ["1. Clients & Related Entities", ""],
    ["   - Clients - Individual client data", ""],
    ["   - Client Beneficiaries - Beneficiary details for individual clients", ""],
    ["   - Client Guardians - Guardian details for minor beneficiaries", ""],
    ["", ""],
    ["2. Corporate Clients & Related Entities", ""],
    ["   - Corporate Clients - Corporate client data", ""],
    ["   - Corporate Directors - Director details for corporate clients", ""],
    ["   - Corporate Beneficiaries - Beneficiary details for corporate clients", ""],
    ["   - Corporate Guardians - Guardian details for minor beneficiaries", ""],
    ["", ""],
    ["3. Agents & Agencies", ""],
    ["   - Agencies - Agency data", ""],
    ["   - Agents - Agent data with bank details", ""],
    ["", ""],
    ["4. Products & Configurations", ""],
    ["   - Product Types - Product type definitions", ""],
    ["   - Products - Detailed product information", ""],
    ["   - Product Commission Config - Commission configurations for products", ""],
    ["   - Product Dividend Schedule - Scheduled dividend payments", ""],
    ["   - Product Agreement Schedule - Scheduled agreement dates", ""],
    ["", ""],
    ["5. Transactions", ""],
    ["   - Product Orders - Product purchase orders", ""],
    ["   - Withdrawals - Client withdrawal transactions", ""],
    ["   - Redemptions - Client redemption transactions", ""],
    ["   - Dividends - Dividend declarations", ""],
    ["   - Dividend Payments - Individual dividend payments to clients", ""],
    ["   - Commissions - Commission payments to agents", ""],
    ["", ""],
    ["General Guidelines:", ""],
    ["- Fields marked as 'Mandatory' must be filled", ""],
    ["- Fields marked as 'Optional' can be left blank", ""],
    ["- Date fields should be in YYYY-MM-DD format", ""],
    ["- For fields with dropdown values, select from the provided options", ""],
    ["- For Boolean fields, use TRUE or FALSE values", ""],
    ["", ""],
    ["Relationship Dependencies:", ""],
    ["- Agencies must be imported before Agents", ""],
    ["- Agents must be imported before Clients and Corporate Clients", ""],
    ["- Product Types must be imported before Products", ""],
    ["- Clients and Products must be imported before Orders", ""],
    ["- Orders must be imported before Withdrawals, Redemptions, and Commissions", ""],
    ["- Products must be imported before Dividend Schedules", ""],
    ["- Dividend Schedules should be imported before Dividends", ""],
    ["- ID fields (client_id, agent_id, etc.) can be left blank for new records and will be auto-generated", ""],
    ["", ""],
    ["Import Process:", ""],
    ["1. Fill in the appropriate sheet with your data", ""],
    ["2. Save the Excel file", ""],
    ["3. Use the admin system to upload and import the data", ""],
    ["4. Review any import errors and fix as needed", ""],
    ["5. Follow the dependency order when importing multiple sheets", ""],
    ["", ""],
    ["Note: This template includes all available fields. Not all fields may be required for your specific import needs.", ""],
]

# Format instructions sheet
for row_idx, row in enumerate(instructions, 1):
    for col_idx, value in enumerate(row, 1):
        ws_instructions.cell(row=row_idx, column=col_idx, value=value)

# Format title
title_cell = ws_instructions['A1']
title_cell.font = Font(bold=True, size=14)

# Make headers and section titles bold
for row in [5, 7, 12, 17, 22, 28, 34, 42, 52]:
    ws_instructions.cell(row=row, column=1).font = Font(bold=True)

# Make subcategories bold
for row in range(8, 11):
    ws_instructions.cell(row=row, column=1).font = Font(bold=True)
for row in range(13, 17):
    ws_instructions.cell(row=row, column=1).font = Font(bold=True)
for row in range(18, 20):
    ws_instructions.cell(row=row, column=1).font = Font(bold=True)
for row in range(23, 28):
    ws_instructions.cell(row=row, column=1).font = Font(bold=True)
for row in range(29, 34):
    ws_instructions.cell(row=row, column=1).font = Font(bold=True)

# Adjust column width
ws_instructions.column_dimensions['A'].width = 80
ws_instructions.column_dimensions['B'].width = 30

# Make instructions sheet active
wb.active = ws_instructions

# Remove the default sheet
wb.remove(default_sheet)

# ====================== BANK REFERENCES SHEET ======================
ws_bank_references = wb.create_sheet(title="Bank References")

# Set up the headers for bank references
bank_ref_headers = ["Bank Name", "SWIFT Code"]
for col_idx, header in enumerate(bank_ref_headers, 1):
    cell = ws_bank_references.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Add bank data to the reference sheet
for row_idx, bank_name in enumerate(bank_names, 2):
    ws_bank_references.cell(row=row_idx, column=1, value=bank_name)
    ws_bank_references.cell(row=row_idx, column=2, value=swift_codes.get(bank_name, ""))
    
# Adjust column widths
ws_bank_references.column_dimensions['A'].width = 40
ws_bank_references.column_dimensions['B'].width = 20

# Add a note at the top explaining the sheet
note_row = ws_bank_references.insert_rows(1)
note_cell = ws_bank_references.cell(row=1, column=1, value="Reference list of banks and SWIFT codes - do not modify this sheet")
note_cell.font = Font(bold=True)
ws_bank_references.merge_cells('A1:B1')
note_cell.alignment = Alignment(horizontal='center')

# ====================== AGENCY SHEET ======================
ws_agency = wb.create_sheet(title="Agencies")

agency_headers = [
    # Basic Information
    "agency_code","agency_id", "agency_name", "agency_reg_number",
    "contact_number", "email_address", "office_address", "office_postcode", 
    "office_city", "office_state", "office_country",
    
    # Leadership Information
    "agency_pic_name", "agency_pic_id",
    
    # Status Information
    "status", "agency_type", "agreement_start_date", "agreement_end_date",
    
    # Bank Account Details
    "bank_name","bank_account_number", "bank_account_holder_name", "bank_address","postcode","city","state","country","bank_swift_code"
]

agency_descriptions = [
    # Basic Information descriptions
    "Agency code (3 characters). Example : ABC",
    "Unique identifier for agency (generated if not provided). Example : ABC0000110015",
    "Full agency name",
    "Agency registration number",
    "Contact phone number",
    "Email address",
    "Office address",
    "Office postal/ZIP code",
    "Office city",
    "Office state/province",
    "Office country",
    
    # Leadership Information descriptions
    "Agency PIC Name",
    "Agency PIC NRIC/Passport number",
    
    # Status Information descriptions
    "Status (TRUE for active, FALSE for inactive)",
    "Agency type (CITADEL or OTHER only)",
    "Agreement start date (YYYY-MM-DD)",
    "Agreement end date (YYYY-MM-DD)",
    
    # Bank Account Details descriptions
    "Bank name",
    "Bank account number",
    "Name as appears on bank account",
    "Bank Address",
    "Bank Postcode",
    "Bank City",
    "Bank State",
    "Bank Country",
    "Bank SWIFT code",
]

mandatory_agency_fields = [
    "agency_code","agency_id","agency_name", "agency_reg_number", "contact_number",
    "email_address", "status", "agency_type", 
    "bank_name","bank_account_number", "bank_account_holder_name", "bank_address","postcode","city","state","country","bank_swift_code"
]

create_headers(ws_agency, agency_headers, agency_descriptions, mandatory_agency_fields)

# Apply validations to their respective columns in the Agency sheet
# add_sheet_specific_validation(ws_agency, "list", '"TRUE,FALSE"', 'P4:P1000', 'status')
# add_sheet_specific_validation(ws_agency, "list", '"CITADEL,OTHER"', 'Q4:Q1000', 'agency_type')
# add_sheet_specific_validation(ws_agency, "list", '"TRUE,FALSE"', 'W4:W1000', 'is_primary_bank_account')  # Updated from U to W due to bank fields


# ====================== AGENTS SHEET ======================
ws_agents = wb.create_sheet(title="Agents")

agent_headers = [
    # Basic Information
    "email", "password",
    "agency_id", "referral_code", "agent_role","recruit_manager_id",

    # Agent Professional Details
    "name","mobile_country_code","mobile_number","identity_card_number", "identity_doc_type",
    "address", "postcode", "city", "state", "country",
    "is_same_corresponding_address",
    "corresponding_address","corresponding_postcode","corresponding_city","corresponding_state","corresponding_country",
    "dob",

    # Bank Account Details
    "bank_name", "bank_account_number","bank_account_holder_name","bank_swift_code",
    "bank_address", "bank_postcode", "bank_city", "bank_state", "bank_country",
]

agent_descriptions = [
    # 1. email
    "Email address used by the agent",
    # 2. password
    "Agent password for authentication (generated default value: PASSWORD if not provided)",
    # 3. agency_name
    "Agency ID the agent belongs to. (Example : CWP1551670001)",
    # 4. referral_code
    "Referral code for the agent (generated if not provided)",
    # 5. agent_role
    "Role of the agent within the agency. (Example : MGR,VP,SVP)",
    # 6. recruit_manager_id
    "ID of the recruiting manager if applicable. (Example : AG001TER0026)",
    # 8. name
    "Full name as it appears on the identity document",
    # 9. mobile_country_code
    "Mobile country code (e.g., +60)",
    # 10. mobile_number
    "Mobile number without the country code and without hypen(-). Example : 1234567890",
    # 11. identity_card_number
    "Identity card/passport number",
    # 12. identity_doc_type
    "Type of identity document (NRIC, PASSPORT, etc.)",
    # 13. address
    "Residential address",
    # 14. postcode
    "Postal/ZIP code of the residential address",
    # 15. city
    "City of the residential address",
    # 16. state
    "State/Province of the residential address",
    # 17. country
    "Country of the residential address",
    # 18. is_same_corresponding_address
    "Indicator if the corresponding address is the same as the residential address (Example : TRUE/FALSE)",
    # 19. corresponding_address
    "Corresponding address if different from residential address",
    # 20. corresponding_postcode
    "Postal/ZIP code of the corresponding address",
    # 21. corresponding_city
    "City of the corresponding address",
    # 22. corresponding_state
    "State/Province of the corresponding address",
    # 23. corresponding_country
    "Country of the corresponding address",
    # 24. dob
    "Date of birth (YYYY-MM-DD)",
    # 25. bank_name
    "Name of the bank for the agent's account",
    # 26. bank_account_number
    "Bank account number for the agent",
    # 27. bank_account_holder_name
    "Name of the account holder (as in bank records)",
    # 28. bank_swift_code
    "SWIFT code of the bank",
    # 29. bank_address
    "Address of the bank",
    # 30. bank_postcode
    "Postal/ZIP code of the bank",
    # 31. bank_city
    "City of the bank",
    # 32. bank_state
    "State/Province of the bank",
    # 33. bank_country
    "Country of the bank"
]


mandatory_agent_fields = [
    "email", "agency_id","agent_role",
    "name","mobile_country_code","mobile_number","identity_card_number", "identity_doc_type",
    "bank_name", "bank_account_number","bank_account_holder_name","bank_swift_code",
]

create_headers(ws_agents, agent_headers, agent_descriptions, mandatory_agent_fields)

# Apply validations to their respective columns in the Agents sheet
# add_sheet_specific_validation(ws_agents, "list", '"ACTIVE,TERMINATED,SUSPENDED"', 'Z4:Z1000', 'status')  # Column shifted from AB to Z
# add_sheet_specific_validation(ws_agents, "list", '"MALE,FEMALE"', 'J4:J1000', 'gender')
# add_sheet_specific_validation(ws_agents, "list", '"NRIC,PASSPORT,BIRTH_CERTIFICATE,OTHER"', 'H4:H1000', 'identity_doc_type')
# add_sheet_specific_validation(ws_agents, "list", '"SINGLE,MARRIED,DIVORCED,WIDOWED"', 'L4:L1000', 'marital_status')
# add_sheet_specific_validation(ws_agents, "list", '"CITIZEN,PERMANENT_RESIDENT,FOREIGNER"', 'M4:M1000', 'residential_status')
# add_sheet_specific_validation(ws_agents, "list", '"SECONDARY,DIPLOMA,BACHELOR,MASTER,PHD"', 'X4:X1000', 'education_level')
# add_sheet_specific_validation(ws_agents, "list", '"TRUE,FALSE"', 'AC4:AC1000', 'is_primary_bank_account')  # Updated from AA to AC due to bank fields
# add_sheet_specific_validation(ws_agents, "textLength", "0", 'D4:D1000', 'mobile_country_code', True)

# ====================== CLIENTS SHEET ======================
ws_clients = wb.create_sheet(title="Clients")

client_headers = [
    "email", "password",
    # Basic Information
    "name",
    "client_id","agent_id",
    "mobile_country_code", "mobile_number", 
    "identity_card_number", "identity_doc_type", 
    "address", "postcode", "city", "state", "country",
    "is_same_corresponding_address", "corresponding_address", "corresponding_postcode", "corresponding_city", "corresponding_state", "corresponding_country",
    "dob", "gender",
    "nationality", "marital_status", "residential_status", 
    "employment_type", "employer_name", "industry_type", "job_title", 
    "employer_address", "employer_postcode", "employer_city", "employer_state", "employer_country",
    "annual_income_declaration", "source_of_income", "source_of_income_remark",

    # PEP Information - Expanded
    "is_pep", "pep_type", "pep_immediate_family_name", "pep_position", "pep_organisation",

    # Banking Details
    "bank_account_number", "bank_account_holder_name", "bank_name", "bank_swift_code",
    "bank_address", "bank_postcode", "bank_city", "bank_state", "bank_country",

    # Agent Reference
    "agent_id", "status",
]

client_descriptions = [
    # 1. email
    "Email address used by the client",
    
    # 2. password
    "Password (or PIN) for client authentication",

    "Name of the client based on the identity card",

    # 3. client_id
    "Unique identifier for the client (auto-generated if not provided)",
    
    # 4. agent_id
    "Identifier of the agent associated with this client",
    
    # 5. mobile_country_code
    "Country calling code for the client's mobile number (e.g., +60)",
    
    # 6. mobile_number
    "Client's mobile number without country code",
    
    # 7. identity_card_number
    "Client's identity card or passport number",
    
    # 8. identity_doc_type
    "Type of identity document (NRIC, PASSPORT, etc.)",
    
    # 9. address
    "Residential address of the client",
    
    # 10. postcode
    "Postal or ZIP code of the residential address",
    
    # 11. city
    "City of the residential address",
    
    # 12. state
    "State or province of the residential address",
    
    # 13. country
    "Country of the residential address",
    
    # 14. is_same_corresponding_address
    "Indicator if the corresponding address is identical to the residential address",
    
    # 15. corresponding_address
    "Corresponding address if different from the residential address",
    
    # 16. corresponding_postcode
    "Postal or ZIP code of the corresponding address",
    
    # 17. corresponding_city
    "City of the corresponding address",
    
    # 18. corresponding_state
    "State or province of the corresponding address",
    
    # 19. corresponding_country
    "Country of the corresponding address",
    
    # 20. dob
    "Date of birth of the client (YYYY-MM-DD)",
    
    # 21. gender
    "Gender of the client (MALE, FEMALE, etc.)",
    
    # 22. nationality
    "Nationality of the client",
    
    # 23. marital_status
    "Marital status (SINGLE, MARRIED, DIVORCED, WIDOWED, etc.)",
    
    # 24. residential_status
    "Residential status (CITIZEN, PERMANENT_RESIDENT, FOREIGNER, etc.)",
    
    # 25. employment_type
    "Client's type of employment (FULL_TIME, PART_TIME, SELF_EMPLOYED, etc.)",
    
    # 26. employer_name
    "Name of the client's current employer",
    
    # 27. industry_type
    "Industry or sector where the client is employed",
    
    # 28. job_title
    "Job title or position the client holds",
    
    # 29. employer_address
    "Physical address of the client's employer",
    
    # 30. employer_postcode
    "Postal or ZIP code of the employer's address",
    
    # 31. employer_city
    "City of the employer's address",
    
    # 32. employer_state
    "State or province of the employer's address",
    
    # 33. employer_country
    "Country of the employer's address",
    
    # 34. annual_income_declaration
    "Client's declared annual income",
    
    # 35. source_of_income
    "Primary source of the client's income",
    
    # 36. source_of_income_remark
    "Additional details regarding the client's source of income",
    
    # 37. is_pep
    "Indicates if the client is a Politically Exposed Person (TRUE/FALSE)",
    
    # 38. pep_type
    "Type of PEP (e.g., DOMESTIC, FOREIGN) if the client is a PEP",
    
    # 39. pep_immediate_family_name
    "Name of the immediate family member who is a PEP, if relevant",
    
    # 40. pep_position
    "Official position or role held by the PEP",
    
    # 41. pep_organisation
    "Name of the organization in which the PEP holds the position",
    
    # 42. bank_account_number
    "Client's bank account number",
    
    # 43. bank_account_holder_name
    "Exact name of the account holder as per bank records",
    
    # 44. bank_name
    "Name of the bank where the account is held",
    
    # 45. bank_swift_code
    "Bank's SWIFT code for international transfers",
    
    # 46. bank_address
    "Physical address of the bank",
    
    # 47. bank_postcode
    "Postal or ZIP code of the bank's address",
    
    # 48. bank_city
    "City where the bank is located",
    
    # 49. bank_state
    "State or province where the bank is located",
    
    # 50. bank_country
    "Country where the bank is located",
    
    # 51. agent_id
    "Reference to the agent overseeing or linked to this client record (duplicate field)",
    
    # 52. status
    "Client's overall status (ACTIVE, SUSPENDED, etc.)"
]

mandatory_client_fields = [
    "email", "agent_id","mobile_country_code", "mobile_number", "name", "identity_card_number", 
    "identity_doc_type", "dob", "gender", "nationality", "marital_status", 
    "residential_status", "address", "postcode", "city", "state", "country",
    "employment_type", "annual_income_declaration", "source_of_income",
    "is_pep", "bank_account_number", "bank_account_holder_name", "bank_name", "bank_swift_code"
]

create_headers(ws_clients, client_headers, client_descriptions, mandatory_client_fields)

# Add dropdowns for enumerated types
# gender_dv = DataValidation(type="list", formula1='"MALE,FEMALE"', allow_blank=True)
# identity_doc_dv = DataValidation(type="list", formula1='"NRIC,PASSPORT"', allow_blank=True)
# marital_status_dv = DataValidation(type="list", formula1='"SINGLE,MARRIED,DIVORCED,WIDOWED"', allow_blank=True)
# residential_status_dv = DataValidation(type="list", formula1='"CITIZEN,PERMANENT_RESIDENT,FOREIGNER"', allow_blank=True)
# employment_type_dv = DataValidation(type="list", formula1='"EMPLOYED,SELF_EMPLOYED,UNEMPLOYED,RETIRED,STUDENT"', allow_blank=True)
# boolean_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
# pep_type_dv = DataValidation(type="list", formula1='"SELF,IMMEDIATE_FAMILY,CLOSE_ASSOCIATE"', allow_blank=True)

# # Create a text data validation for mobile_country_code to prevent any dropdown from being applied
# mobile_country_code_dv = DataValidation(type="textLength", operator="greaterThan", formula1="0", allow_blank=True)

# # Apply validations to their respective columns in the Clients sheet
# add_validation_to_sheet(ws_clients, mobile_country_code_dv, 'D4:D1000', 'mobile_country_code')
# add_validation_to_sheet(ws_clients, gender_dv, 'J4:J1000', 'gender')
# add_validation_to_sheet(ws_clients, identity_doc_dv, 'H4:H1000', 'identity_doc_type')
# add_validation_to_sheet(ws_clients, marital_status_dv, 'L4:L1000', 'marital_status')
# add_validation_to_sheet(ws_clients, residential_status_dv, 'M4:M1000', 'residential_status')
# add_validation_to_sheet(ws_clients, employment_type_dv, 'Y4:Y1000', 'employment_type')
# add_validation_to_sheet(ws_clients, pep_type_dv, 'AL4:AL1000', 'pep_type')

# # Apply boolean validations to multiple columns
# boolean_validations = {
#     'S4:S1000': 'is_same_corresponding_address',
#     'AJ4:AJ1000': 'is_pep',
#     'AQ4:AQ1000': 'is_primary_bank_account',  # Updated to AQ since bank_name and swift_code are added back
#     'AR4:AR1000': 'status',  # Updated to AR
#     'AS4:AS1000': 'has_beneficiary'  # Updated to AS
# }

# for cell_range, field_name in boolean_validations.items():
#     add_validation_to_sheet(ws_clients, boolean_dv, cell_range, field_name)

# ====================== CLIENT BENEFICIARIES SHEET ======================
ws_beneficiaries = wb.create_sheet(title="Client Beneficiaries")

beneficiary_headers = [
    "client_id", "beneficiary_id", "name", "identity_card_number", "identity_doc_type",
    "dob", "gender", "nationality", "relationship", "address", "postcode", "city", "state", "country",
    "mobile_country_code", "mobile_number", "email", "allocation_percentage", "priority_order",
    "is_minor", "has_guardian"
]

beneficiary_descriptions = [
    "Client ID to associate with this beneficiary",
    "Unique identifier for beneficiary (generated if not provided)",
    "Full name of beneficiary as in ID",
    "Identity card/passport number",
    "Type of identity document (NRIC, PASSPORT, etc.)",
    "Date of birth (YYYY-MM-DD)",
    "Gender (MALE, FEMALE)",
    "Nationality",
    "Relationship to client (SPOUSE, CHILD, PARENT, SIBLING, OTHER)",
    "Beneficiary address",
    "Postal/ZIP code",
    "City",
    "State/Province",
    "Country",
    "Mobile country code (e.g., +60)",
    "Mobile number without country code",
    "Email address",
    "Allocation percentage (must sum to 100% per client)",
    "Priority order number",
    "Is beneficiary a minor (TRUE/FALSE)",
    "Does beneficiary have a guardian (TRUE/FALSE)"
]

mandatory_beneficiary_fields = [
    "client_id", "name", "identity_card_number", "identity_doc_type", "dob", "relationship",
    "allocation_percentage", "is_minor", "has_guardian"
]

create_headers(ws_beneficiaries, beneficiary_headers, beneficiary_descriptions, mandatory_beneficiary_fields)

# Apply validations to their respective columns in the Client Beneficiaries sheet
add_sheet_specific_validation(ws_beneficiaries, "list", '"MALE,FEMALE"', 'G4:G1000', 'gender')
add_sheet_specific_validation(ws_beneficiaries, "list", '"NRIC,PASSPORT,BIRTH_CERTIFICATE,OTHER"', 'E4:E1000', 'identity_doc_type')
add_sheet_specific_validation(ws_beneficiaries, "list", '"SELF,FAMILY,ASSOCIATE,CHILD,GUARDIAN,PARENTS,PARTNER,SPOUSE,GRANDPARENT,FRIEND,FIANCE,SIBLING,NIECE,NEPHEW,GRAND_DAUGHTER,GRAND_SON,GOD_PARENT,MOTHER_IN_LAW,FATHER_IN_LAW,SON_IN_LAW,DAUGHTER_IN_LAW"', 'I4:I1000', 'relationship')
add_sheet_specific_validation(ws_beneficiaries, "textLength", "0", 'O4:O1000', 'mobile_country_code', True)
add_sheet_specific_validation(ws_beneficiaries, "list", '"TRUE,FALSE"', 'T4:T1000', 'is_minor')
add_sheet_specific_validation(ws_beneficiaries, "list", '"TRUE,FALSE"', 'U4:U1000', 'has_guardian')

# ====================== CLIENT GUARDIANS SHEET ======================
ws_guardians = wb.create_sheet(title="Client Guardians")

guardian_headers = [
    "beneficiary_id", "guardian_id", "name", "identity_card_number", "identity_doc_type",
    "dob", "gender", "nationality", "relationship", "address", "postcode", "city", "state", "country",
    "mobile_country_code", "mobile_number", "email", "guardian_start_date", "guardian_end_date"
]

guardian_descriptions = [
    "Beneficiary ID to associate with this guardian",
    "Unique identifier for guardian (generated if not provided)",
    "Full name of guardian as in ID",
    "Identity card/passport number",
    "Type of identity document (NRIC, PASSPORT, etc.)",
    "Date of birth (YYYY-MM-DD)",
    "Gender (MALE, FEMALE)",
    "Nationality",
    "Relationship to beneficiary (PARENT, LEGAL_GUARDIAN, OTHER)",
    "Guardian address",
    "Postal/ZIP code",
    "City",
    "State/Province",
    "Country",
    "Mobile country code (e.g., +60)",
    "Mobile number without country code",
    "Email address",
    "Date guardianship starts (YYYY-MM-DD)",
    "Date guardianship ends (YYYY-MM-DD, if applicable)"
]

mandatory_guardian_fields = [
    "beneficiary_id", "name", "identity_card_number", "identity_doc_type", "dob", "relationship",
    "mobile_country_code", "mobile_number", "guardian_start_date"
]

create_headers(ws_guardians, guardian_headers, guardian_descriptions, mandatory_guardian_fields)

# Apply validations to their respective columns in the Client Guardians sheet
add_sheet_specific_validation(ws_guardians, "list", '"MALE,FEMALE"', 'G4:G1000', 'gender')
add_sheet_specific_validation(ws_guardians, "list", '"NRIC,PASSPORT,BIRTH_CERTIFICATE,OTHER"', 'E4:E1000', 'identity_doc_type')
add_sheet_specific_validation(ws_guardians, "list", '"SELF,FAMILY,ASSOCIATE,CHILD,GUARDIAN,PARENTS,PARTNER,SPOUSE,GRANDPARENT,FRIEND,FIANCE,SIBLING,NIECE,NEPHEW,GRAND_DAUGHTER,GRAND_SON,GOD_PARENT,MOTHER_IN_LAW,FATHER_IN_LAW,SON_IN_LAW,DAUGHTER_IN_LAW"', 'I4:I1000', 'relationship')
add_sheet_specific_validation(ws_guardians, "textLength", "0", 'O4:O1000', 'mobile_country_code', True)


# ====================== CORPORATE CLIENTS SHEET ======================
ws_corporate = wb.create_sheet(title="Corporate Clients")

corporate_headers = [
    # Authorizer details
    "authorizer_email", "authorizer_mobile_country_code", "authorizer_mobile_number", 
    "authorizer_name", "authorizer_identity_card_number", "authorizer_identity_doc_type", 
    "authorizer_dob", "authorizer_gender", "authorizer_nationality",
    "authorizer_address", "authorizer_postcode", "authorizer_city", "authorizer_state", "authorizer_country",
    "authorizer_position", "authorizer_appointment_date",
    
    # Corporate details
    "corporate_client_id", "entity_name", "entity_type", "registration_number", 
    "date_incorporation", "place_incorporation", "business_type",
    "registered_address", "postcode", "city", "state", "country",
    "is_different_business_address", "business_address", "business_postcode", "business_city", "business_state", "business_country",
    
    # Contact Information
    "contact_is_myself", "contact_name", "contact_designation", "contact_mobile_country_code", "contact_mobile_number", "contact_email",
    
    # Corporate Wealth and Income
    "annual_revenue", "number_of_employees", "asset_value", "liability_value", 
    "source_of_income", "source_of_income_remark", "business_sector", "business_activity",
    
    # PEP Information
    "is_related_to_pep", "pep_relationship_description",
    
    # Miscellaneous
    "reference_number", "approval_status", "approval_status_remark",
    "agent_id",
    
    # Beneficiary Information
    "has_beneficiary", "beneficiary_count"
]

corporate_descriptions = [
    # Authorizer descriptions
    "Email address of authorized representative",
    "Mobile country code of representative (e.g., +60)",
    "Mobile number of representative without country code",
    "Full name of authorized representative as in ID",
    "Identity card/passport number of representative",
    "Type of identity document for representative (NRIC, PASSPORT, etc.)",
    "Date of birth of representative (YYYY-MM-DD)",
    "Gender of representative (MALE, FEMALE)",
    "Nationality of representative",
    "Address of representative",
    "Postal/ZIP code of representative",
    "City of representative",
    "State/Province of representative",
    "Country of representative",
    "Position of authorizer in company (DIRECTOR, CEO, CFO, etc.)",
    "Date of appointment (YYYY-MM-DD)",
    
    # Corporate details descriptions
    "Unique identifier for corporate client (generated if not provided)",
    "Entity/Company name",
    "Entity type (SDN BHD, BHD, etc.)",
    "Company registration number",
    "Date of incorporation (YYYY-MM-DD)",
    "Place of incorporation",
    "Type of business",
    "Registered company address",
    "Company postal/ZIP code",
    "Company city",
    "Company state/province",
    "Company country",
    "Is business address different from registered address (TRUE/FALSE)",
    "Business address if different",
    "Business postal/ZIP code if different",
    "Business city if different",
    "Business state/province if different",
    "Business country if different",
    
    # Contact Information descriptions
    "Is contact person same as representative (TRUE/FALSE)",
    "Contact person name if different",
    "Designation of contact person",
    "Contact person mobile country code",
    "Contact person mobile number",
    "Contact person email",
    
    # Corporate Wealth and Income descriptions
    "Annual revenue (in local currency)",
    "Number of employees",
    "Total asset value (in local currency)",
    "Total liability value (in local currency)",
    "Source of corporate income",
    "Additional remarks for source of income",
    "Business sector (MANUFACTURING, SERVICES, RETAIL, etc.)",
    "Detailed business activity description",
    
    # PEP Information descriptions
    "Is the corporate entity related to any PEP (TRUE/FALSE)",
    "Description of PEP relationship if applicable",
    
    # Miscellaneous descriptions
    "Reference number",
    "Approval status (IN_REVIEW, APPROVED, REJECTED, DRAFT)",
    "Approval remark if applicable",
    "Agent ID of associated agent",
    
    # Beneficiary Information descriptions
    "Has beneficiary designated (TRUE/FALSE)",
    "Number of beneficiaries (if applicable)"
]

mandatory_corporate_fields = [
    # Required Authorizer Fields
    "authorizer_email", "authorizer_mobile_country_code", "authorizer_mobile_number", 
    "authorizer_name", "authorizer_identity_card_number", "authorizer_identity_doc_type", 
    "authorizer_position",
    
    # Required Corporate Fields
    "entity_name", "entity_type", "registration_number", "date_incorporation", 
    "place_incorporation", "business_type", "registered_address", "postcode", 
    "city", "state", "country", "is_different_business_address",
    "contact_is_myself", 
    
    # Required Financial Fields
    "annual_revenue", "source_of_income", "business_sector",
    
    # Required Bank Fields
    "bank_account_number", "bank_account_holder_name",
    
    # Required Status Fields
    "approval_status", "is_related_to_pep", "has_beneficiary"
]

create_headers(ws_corporate, corporate_headers, corporate_descriptions, mandatory_corporate_fields)

# Apply validations to their respective columns in the Corporate Clients sheet
add_sheet_specific_validation(ws_corporate, "list", '"MALE,FEMALE"', 'H4:H1000', 'gender')
add_sheet_specific_validation(ws_corporate, "list", '"NRIC,PASSPORT,BIRTH_CERTIFICATE,OTHER"', 'F4:F1000', 'identity_doc_type')
add_sheet_specific_validation(ws_corporate, "list", '"DIRECTOR,CEO,CFO,COO,SECRETARY,SHAREHOLDER,OTHER"', 'O4:O1000', 'authorizer_position')
add_sheet_specific_validation(ws_corporate, "list", '"MANUFACTURING,SERVICES,RETAIL,FINANCE,TECHNOLOGY,AGRICULTURE,CONSTRUCTION,EDUCATION,HEALTHCARE,OTHER"', 'AM4:AM1000', 'business_sector')
add_sheet_specific_validation(ws_corporate, "list", '"IN_REVIEW,APPROVED,REJECTED,DRAFT"', 'AZ4:AZ1000', 'approval_status')
add_sheet_specific_validation(ws_corporate, "list", '"SDN BHD,BHD,ENTERPRISE,PARTNERSHIP,LIMITED LIABILITY PARTNERSHIP,OTHER"', 'R4:R1000', 'entity_type')
add_sheet_specific_validation(ws_corporate, "textLength", "0", 'B4:B1000', 'authorizer_mobile_country_code', True)
add_sheet_specific_validation(ws_corporate, "textLength", "0", 'T4:T1000', 'contact_mobile_country_code', True)

# Apply boolean validations to multiple fields
boolean_fields = {
    'M4:M1000': 'is_different_business_address',
    'S4:S1000': 'contact_is_myself',
    'AO4:AO1000': 'is_related_to_pep',
    'AP4:AP1000': 'is_primary_bank_account',  # Updated from AQ to AP since bank_name and swift_code are removed
    'AT4:AT1000': 'has_beneficiary'  # Updated from AV to AT
}

for cell_range, field_name in boolean_fields.items():
    add_sheet_specific_validation(ws_corporate, "list", '"TRUE,FALSE"', cell_range, field_name)

# ====================== CORPORATE DIRECTORS SHEET ======================
ws_directors = wb.create_sheet(title="Corporate Directors")

director_headers = [
    "corporate_client_id", "director_id", "name", "identity_card_number", "identity_doc_type",
    "dob", "gender", "nationality", "position", "appointment_date", "termination_date", 
    "address", "postcode", "city", "state", "country",
    "mobile_country_code", "mobile_number", "email", 
    "shareholding_percentage", "is_authorizer", "is_pep", "pep_details"
]

director_descriptions = [
    "Corporate Client ID to associate with this director",
    "Unique identifier for director (generated if not provided)",
    "Full name of director as in ID",
    "Identity card/passport number",
    "Type of identity document (NRIC, PASSPORT, etc.)",
    "Date of birth (YYYY-MM-DD)",
    "Gender (MALE, FEMALE)",
    "Nationality",
    "Position in the company (DIRECTOR, CHAIRMAN, etc.)",
    "Date of appointment (YYYY-MM-DD)",
    "Date of termination if applicable (YYYY-MM-DD)",
    "Director address",
    "Postal/ZIP code",
    "City",
    "State/Province",
    "Country",
    "Mobile country code (e.g., +60)",
    "Mobile number without country code",
    "Email address",
    "Shareholding percentage if applicable",
    "Is this person the same as the authorized representative (TRUE/FALSE)",
    "Is director a politically exposed person (TRUE/FALSE)",
    "PEP details if applicable"
]

mandatory_director_fields = [
    "corporate_client_id", "name", "identity_card_number", "identity_doc_type", 
    "nationality", "position", "appointment_date", "shareholding_percentage", 
    "is_authorizer", "is_pep"
]

create_headers(ws_directors, director_headers, director_descriptions, mandatory_director_fields)

# Apply validations to their respective columns in the Corporate Directors sheet
add_sheet_specific_validation(ws_directors, "list", '"MALE,FEMALE"', 'G4:G1000', 'gender')
add_sheet_specific_validation(ws_directors, "list", '"NRIC,PASSPORT,BIRTH_CERTIFICATE,OTHER"', 'E4:E1000', 'identity_doc_type')
add_sheet_specific_validation(ws_directors, "list", '"DIRECTOR,CHAIRMAN,EXECUTIVE_DIRECTOR,NON_EXECUTIVE_DIRECTOR,INDEPENDENT_DIRECTOR,OTHER"', 'I4:I1000', 'position')
add_sheet_specific_validation(ws_directors, "list", '"TRUE,FALSE"', 'U4:U1000', 'is_authorizer')
add_sheet_specific_validation(ws_directors, "list", '"TRUE,FALSE"', 'V4:V1000', 'is_pep')
add_sheet_specific_validation(ws_directors, "textLength", "0", 'Q4:Q1000', 'mobile_country_code', True)

# ====================== CORPORATE BENEFICIARIES SHEET ======================
ws_corporate_beneficiaries = wb.create_sheet(title="Corporate Beneficiaries")

corporate_beneficiary_headers = [
    "corporate_client_id", "beneficiary_id", "beneficiary_type", 
    # If Individual
    "name", "identity_card_number", "identity_doc_type", "dob", "gender", "nationality", 
    "relationship", "address", "postcode", "city", "state", "country",
    "mobile_country_code", "mobile_number", "email", 
    # If Corporate
    "company_name", "company_registration_number", "company_address", "company_postcode", 
    "company_city", "company_state", "company_country", "company_contact_person",
    # Common
    "allocation_percentage", "priority_order", "is_minor", "has_guardian"
]

corporate_beneficiary_descriptions = [
    "Corporate Client ID to associate with this beneficiary",
    "Unique identifier for beneficiary (generated if not provided)",
    "Type of beneficiary (INDIVIDUAL, CORPORATE)",
    # Individual descriptions
    "Full name of beneficiary as in ID (if individual)",
    "Identity card/passport number (if individual)",
    "Type of identity document (NRIC, PASSPORT, etc.) (if individual)",
    "Date of birth (YYYY-MM-DD) (if individual)",
    "Gender (MALE, FEMALE) (if individual)",
    "Nationality (if individual)",
    "Relationship to corporate client (SHAREHOLDER, DIRECTOR, OTHER) (if individual)",
    "Beneficiary address (if individual)",
    "Postal/ZIP code (if individual)",
    "City (if individual)",
    "State/Province (if individual)",
    "Country (if individual)",
    "Mobile country code (e.g., +60) (if individual)",
    "Mobile number without country code (if individual)",
    "Email address (if individual)",
    # Corporate descriptions
    "Company name (if corporate beneficiary)",
    "Registration number (if corporate beneficiary)",
    "Company address (if corporate beneficiary)",
    "Company postal/ZIP code (if corporate beneficiary)",
    "Company city (if corporate beneficiary)",
    "Company state/province (if corporate beneficiary)",
    "Company country (if corporate beneficiary)",
    "Contact person at beneficiary company (if corporate beneficiary)",
    # Common descriptions
    "Allocation percentage (must sum to 100% per corporate client)",
    "Priority order number",
    "Is beneficiary a minor (TRUE/FALSE, for individuals only)",
    "Does beneficiary have a guardian (TRUE/FALSE, for minor beneficiaries only)"
]

mandatory_corporate_beneficiary_fields = [
    "corporate_client_id", "beneficiary_type", "allocation_percentage"
]

create_headers(ws_corporate_beneficiaries, corporate_beneficiary_headers, corporate_beneficiary_descriptions, mandatory_corporate_beneficiary_fields)

# Apply validations to their respective columns in the Corporate Beneficiaries sheet
add_sheet_specific_validation(ws_corporate_beneficiaries, "list", '"INDIVIDUAL,CORPORATE"', 'C4:C1000', 'beneficiary_type')
add_sheet_specific_validation(ws_corporate_beneficiaries, "list", '"MALE,FEMALE"', 'H4:H1000', 'gender')
add_sheet_specific_validation(ws_corporate_beneficiaries, "list", '"NRIC,PASSPORT,BIRTH_CERTIFICATE,OTHER"', 'F4:F1000', 'identity_doc_type')
add_sheet_specific_validation(ws_corporate_beneficiaries, "list", '"SELF,FAMILY,ASSOCIATE,CHILD,GUARDIAN,PARENTS,PARTNER,SPOUSE,GRANDPARENT,FRIEND,FIANCE,SIBLING,NIECE,NEPHEW,GRAND_DAUGHTER,GRAND_SON,GOD_PARENT,MOTHER_IN_LAW,FATHER_IN_LAW,SON_IN_LAW,DAUGHTER_IN_LAW"', 'J4:J1000', 'relationship')
add_sheet_specific_validation(ws_corporate_beneficiaries, "textLength", "0", 'P4:P1000', 'mobile_country_code', True)
add_sheet_specific_validation(ws_corporate_beneficiaries, "list", '"TRUE,FALSE"', 'AB4:AB1000', 'is_minor')
add_sheet_specific_validation(ws_corporate_beneficiaries, "list", '"TRUE,FALSE"', 'AC4:AC1000', 'has_guardian')

# ====================== CORPORATE GUARDIANS SHEET ======================
ws_corporate_guardians = wb.create_sheet(title="Corporate Guardians")

corporate_guardian_headers = [
    "beneficiary_id", "guardian_id", "name", "identity_card_number", "identity_doc_type",
    "dob", "gender", "nationality", "relationship", "address", "postcode", "city", "state", "country",
    "mobile_country_code", "mobile_number", "email", "guardian_start_date", "guardian_end_date"
]

corporate_guardian_descriptions = [
    "Beneficiary ID to associate with this guardian",
    "Unique identifier for guardian (generated if not provided)",
    "Full name of guardian as in ID",
    "Identity card/passport number",
    "Type of identity document (NRIC, PASSPORT, etc.)",
    "Date of birth (YYYY-MM-DD)",
    "Gender (MALE, FEMALE)",
    "Nationality",
    "Relationship to beneficiary (PARENT, LEGAL_GUARDIAN, OTHER)",
    "Guardian address",
    "Postal/ZIP code",
    "City",
    "State/Province",
    "Country",
    "Mobile country code (e.g., +60)",
    "Mobile number without country code",
    "Email address",
    "Date guardianship starts (YYYY-MM-DD)",
    "Date guardianship ends (YYYY-MM-DD, if applicable)"
]

mandatory_corporate_guardian_fields = [
    "beneficiary_id", "name", "identity_card_number", "identity_doc_type", "dob", "relationship",
    "mobile_country_code", "mobile_number", "guardian_start_date"
]

create_headers(ws_corporate_guardians, corporate_guardian_headers, corporate_guardian_descriptions, mandatory_corporate_guardian_fields)

# Apply validations to their respective columns in the Corporate Guardians sheet
add_sheet_specific_validation(ws_corporate_guardians, "list", '"MALE,FEMALE"', 'G4:G1000', 'gender')
add_sheet_specific_validation(ws_corporate_guardians, "list", '"NRIC,PASSPORT,BIRTH_CERTIFICATE,OTHER"', 'E4:E1000', 'identity_doc_type')
add_sheet_specific_validation(ws_corporate_guardians, "list", '"SELF,FAMILY,ASSOCIATE,CHILD,GUARDIAN,PARENTS,PARTNER,SPOUSE,GRANDPARENT,FRIEND,FIANCE,SIBLING,NIECE,NEPHEW,GRAND_DAUGHTER,GRAND_SON,GOD_PARENT,MOTHER_IN_LAW,FATHER_IN_LAW,SON_IN_LAW,DAUGHTER_IN_LAW"', 'I4:I1000', 'relationship')
add_sheet_specific_validation(ws_corporate_guardians, "textLength", "0", 'O4:O1000', 'mobile_country_code', True)

# ====================== PRODUCT TYPE SHEET ======================
# ws_product_types = wb.create_sheet(title="Product Types")

# product_type_headers = [
#     "type_name","status"
# ]

# product_type_descriptions = [
#     "Name of the product type",
#     "Is product type active (TRUE/FALSE)"
# ]

# mandatory_product_type_fields = [
# "type_name","status"
# ]

# create_headers(ws_product_types, product_type_headers, product_type_descriptions, mandatory_product_type_fields)

# # ====================== PRODUCTS SHEET ======================
# ws_products = wb.create_sheet(title="Products")

# product_headers = [
#     # Basic Information
#     "product_type_name", "product_name", "product_code", 
#     "tranche_size", "minimum_investment", "is_prorated", "lock_in_period",
#     "bank_name", "incremental_amount", "investment_tenure_month", "status", "publish"
# ]

# product_descriptions = [
#     # Basic Information descriptions
#     "Product type name from Product Type Sheet",
#     "Full name of the product",
#     "Short code/identifier for the product",
#     "Available tranche size for the product",
#     "Minimum investment amount",
#     "Is this product prorated (TRUE/FALSE)",
#     "Lock-in period in months",
#     "Bank name for the product",
#     "Incremental investment amount",
#     "Investment tenure in months",
#     "Product status (TRUE/FALSE) True if the product is available for sale",
#     "Is this product published (TRUE/FALSE) True if the product is to displayed in citadel app"
# ]

# mandatory_product_fields = [
#     "product_type_name", "product_name", "product_code", 
#     "tranche_size", "minimum_investment", "is_prorated", "lock_in_period",
#     "bank_name", "incremental_amount", "investment_tenure_month", "status", "publish"
# ]

# create_headers(ws_products, product_headers, product_descriptions, mandatory_product_fields)

# return_type_dv = DataValidation(type="list", formula1='"FIXED,VARIABLE,HYBRID"', allow_blank=True)
# # frequency_dv = DataValidation(type="list", formula1='"MONTHLY,QUARTERLY,SEMI_ANNUALLY,ANNUALLY,IRREGULAR"', allow_blank=True)
# # product_status_dv = DataValidation(type="list", formula1='"ACTIVE,INACTIVE,PENDING,SOLD_OUT,MATURED"', allow_blank=True)
# # currency_dv = DataValidation(type="list", formula1='"MYR,USD,SGD,EUR,GBP,JPY,CNY,AUD,OTHER"', allow_blank=True)

# # # Apply validations to their respective columns in the Products sheet
# # add_sheet_specific_validation(ws_products, "list", '"FIXED,VARIABLE,HYBRID"', 'O4:O1000', 'return_type')
# # add_sheet_specific_validation(ws_products, "list", '"MONTHLY,QUARTERLY,SEMI_ANNUALLY,ANNUALLY,IRREGULAR"', 'P4:P1000', 'return_payment_frequency')
# # add_sheet_specific_validation(ws_products, "list", '"MONTHLY,QUARTERLY,SEMI_ANNUALLY,ANNUALLY,IRREGULAR"', 'X4:X1000', 'agreement_schedule_frequency')
# # add_sheet_specific_validation(ws_products, "list", '"MONTHLY,QUARTERLY,SEMI_ANNUALLY,ANNUALLY,IRREGULAR"', 'Y4:Y1000', 'dividend_payment_frequency')
# # add_sheet_specific_validation(ws_products, "list", '"ACTIVE,INACTIVE,PENDING,SOLD_OUT,MATURED"', 'Z4:Z1000', 'status')
# # add_sheet_specific_validation(ws_products, "list", '"MYR,USD,SGD,EUR,GBP,JPY,CNY,AUD,OTHER"', 'F4:F1000', 'currency')
# # add_sheet_specific_validation(ws_products, "list", '"TRUE,FALSE"', 'AA4:AA1000', 'is_featured')
# # add_sheet_specific_validation(ws_products, "list", '"TRUE,FALSE"', 'AB4:AB1000', 'is_shariah_compliant')

# # ====================== PRODUCT COMMISSION CONFIG SHEET ======================
# ws_commission_config = wb.create_sheet(title="Product Commission Config")

# commission_config_headers = [
#     "commission_config_id", "product_id", "commission_type", "is_agency_commission",
#     "tier_level", "start_value", "end_value", "rate_percentage", "is_active",
#     "effective_date", "expiry_date"
# ]

# commission_config_descriptions = [
#     "Unique identifier for commission config (generated if not provided)",
#     "ID of the product this commission config applies to",
#     "Type of commission (SALES, RECRUITMENT, OVERRIDE, BONUS, etc.)",
#     "Is this for agency commission (TRUE) or in-house commission (FALSE)",
#     "Tier level for commission rate (1, 2, 3, etc.)",
#     "Starting value for this commission tier",
#     "Ending value for this commission tier",
#     "Commission rate percentage for this tier",
#     "Is this commission config active (TRUE/FALSE)",
#     "Effective date for commission config (YYYY-MM-DD)",
#     "Expiry date for commission config if applicable (YYYY-MM-DD)"
# ]

# mandatory_commission_config_fields = [
#     "product_id", "commission_type", "is_agency_commission", 
#     "rate_percentage", "is_active", "effective_date"
# ]

# create_headers(ws_commission_config, commission_config_headers, commission_config_descriptions, mandatory_commission_config_fields)

# commission_type_dv = DataValidation(type="list", formula1='"SALES,RECRUITMENT,OVERRIDE,BONUS,MANAGEMENT,RENEWAL"', allow_blank=True)

# # Apply validations to their respective columns in the Product Commission Config sheet
# add_sheet_specific_validation(ws_commission_config, "list", '"SALES,RECRUITMENT,OVERRIDE,BONUS,MANAGEMENT,RENEWAL"', 'C4:C1000', 'commission_type')
# add_sheet_specific_validation(ws_commission_config, "list", '"TRUE,FALSE"', 'D4:D1000', 'is_agency_commission')
# add_sheet_specific_validation(ws_commission_config, "list", '"TRUE,FALSE"', 'I4:I1000', 'is_active')

# # ====================== PRODUCT DIVIDEND SCHEDULE SHEET ======================
# ws_dividend_schedule = wb.create_sheet(title="Product Dividend Schedule")

# dividend_schedule_headers = [
#     "schedule_id", "product_id", "payment_date", "record_date", "payment_amount",
#     "payment_percentage", "payment_status", "remarks"
# ]

# dividend_schedule_descriptions = [
#     "Unique identifier for dividend schedule (generated if not provided)",
#     "ID of the product this dividend schedule applies to",
#     "Date of dividend payment (YYYY-MM-DD)",
#     "Record date for eligibility (YYYY-MM-DD)",
#     "Payment amount per unit if fixed",
#     "Payment percentage if variable",
#     "Payment status (SCHEDULED, PAID, CANCELLED)",
#     "Additional remarks or notes"
# ]

# mandatory_dividend_schedule_fields = [
#     "product_id", "payment_date", "record_date", "payment_status"
# ]

# create_headers(ws_dividend_schedule, dividend_schedule_headers, dividend_schedule_descriptions, mandatory_dividend_schedule_fields)

# payment_status_dv = DataValidation(type="list", formula1='"SCHEDULED,PAID,CANCELLED"', allow_blank=True)

# # Apply validations to their respective columns in the Product Dividend Schedule sheet
# # add_validation_to_sheet(ws_dividend_schedule, payment_status_dv, 'G4:G1000', 'payment_status')
# # payment_status_dv.add(f'G4:G1000')
# add_sheet_specific_validation(ws_dividend_schedule, "list", '"SCHEDULED,PAID,CANCELLED"', 'G4:G1000', 'payment_status')

# # ====================== PRODUCT AGREEMENT SCHEDULE SHEET ======================
# ws_agreement_schedule = wb.create_sheet(title="Product Agreement Schedule")

# agreement_schedule_headers = [
#     "schedule_id", "product_id", "agreement_date", "agreement_details",
#     "is_executed", "executed_date", "executed_by", "remarks"
# ]

# agreement_schedule_descriptions = [
#     "Unique identifier for agreement schedule (generated if not provided)",
#     "ID of the product this agreement schedule applies to",
#     "Date of scheduled agreement (YYYY-MM-DD)",
#     "Details or type of agreement",
#     "Has the agreement been executed (TRUE/FALSE)",
#     "Date the agreement was executed (YYYY-MM-DD)",
#     "User ID or name who executed the agreement",
#     "Additional remarks or notes"
# ]

# mandatory_agreement_schedule_fields = [
#     "product_id", "agreement_date", "agreement_details"
# ]

# create_headers(ws_agreement_schedule, agreement_schedule_headers, agreement_schedule_descriptions, mandatory_agreement_schedule_fields)

# # Apply validations to their respective columns in the Product Agreement Schedule sheet
# # add_validation_to_sheet(ws_agreement_schedule, boolean_dv, 'E4:E1000', 'is_executed')

# # ws_agreement_schedule.add_data_validation(boolean_dv)
# # boolean_dv.add(f'E4:E1000')  # is_executed
# add_sheet_specific_validation(ws_agreement_schedule, "list", '"TRUE,FALSE"', 'E4:E1000', 'is_executed')

# # ====================== PRODUCT ORDERS SHEET ======================
# ws_product_orders = wb.create_sheet(title="Product Orders")

# product_order_headers = [
#     # Order Information
#     "order_id", "order_reference", "order_date", "client_id", "client_type",
#     "product_id", "order_amount", "order_units", "unit_price",
#     "payment_method", "payment_status", "order_status",
    
#     # Transaction Information
#     "transaction_date", "transaction_reference", 
#     "bank_account_number", "account_holder_name",
#     "payment_proof_url", "verified_by", "verification_date",
    
#     # Order Metadata
#     "agent_id", "agency_id", "sales_channel", "promotion_code",
#     "order_notes", "created_by", "approval_status", "approval_date", "approved_by",
    
#     # Beneficiary Information
#     "has_beneficiaries",
    
#     # Primary Beneficiary 1
#     "beneficiary1_id", "beneficiary1_allocation_percentage",
#     "beneficiary1_name", "beneficiary1_identity_card_number", "beneficiary1_relationship",
#     "beneficiary1_is_minor", "beneficiary1_has_guardian",
    
#     # Primary Beneficiary 2
#     "beneficiary2_id", "beneficiary2_allocation_percentage",
#     "beneficiary2_name", "beneficiary2_identity_card_number", "beneficiary2_relationship",
#     "beneficiary2_is_minor", "beneficiary2_has_guardian",
    
#     # Primary Beneficiary 3
#     "beneficiary3_id", "beneficiary3_allocation_percentage",
#     "beneficiary3_name", "beneficiary3_identity_card_number", "beneficiary3_relationship",
#     "beneficiary3_is_minor", "beneficiary3_has_guardian",
    
#     # Guardian Information (for minor beneficiaries)
#     "guardian1_id", "guardian1_name", "guardian1_identity_card_number", "guardian1_relationship", "guardian1_for_beneficiary",
#     "guardian2_id", "guardian2_name", "guardian2_identity_card_number", "guardian2_relationship", "guardian2_for_beneficiary"
# ]

# product_order_descriptions = [
#     # Order Information descriptions
#     "Unique identifier for order (generated if not provided)",
#     "Order reference number",
#     "Date order was placed (YYYY-MM-DD)",
#     "ID of client who placed the order",
#     "Type of client (INDIVIDUAL, CORPORATE)",
#     "ID of product being ordered",
#     "Order amount in product's currency",
#     "Number of units ordered",
#     "Unit price at time of order",
#     "Payment method (BANK_TRANSFER, CREDIT_CARD, CHEQUE, etc.)",
#     "Payment status (PENDING, PAID, FAILED, REFUNDED)",
#     "Order status (PENDING, CONFIRMED, CANCELLED, COMPLETED)",
    
#     # Transaction Information descriptions
#     "Date transaction was completed (YYYY-MM-DD)",
#     "Bank transaction reference",
#     "Bank account number for dividend payout",
#     "Name of account holder for dividend payout",
#     "URL to payment proof document",
#     "User ID/name who verified the payment",
#     "Date payment was verified (YYYY-MM-DD)",
    
#     # Order Metadata descriptions
#     "ID of agent who facilitated the order",
#     "ID of agency the agent belongs to",
#     "Sales channel (DIRECT, AGENT, ONLINE, PHONE, etc.)",
#     "Promotion code if applicable",
#     "Additional notes or remarks",
#     "User ID/name who created the order",
#     "Approval status (PENDING, APPROVED, REJECTED)",
#     "Date of approval/rejection (YYYY-MM-DD)",
#     "User ID/name who approved/rejected the order",
    
#     # Beneficiary Information descriptions
#     "Has beneficiaries designated (TRUE/FALSE)",
    
#     # Primary Beneficiary 1 descriptions
#     "ID of first beneficiary (must exist in beneficiary table)",
#     "Allocation percentage for first beneficiary",
#     "Name of first beneficiary",
#     "Identity card/passport number of first beneficiary",
#     "Relationship to client (SPOUSE, CHILD, PARENT, etc.)",
#     "Is beneficiary a minor (TRUE/FALSE)",
#     "Does beneficiary have a guardian (TRUE/FALSE)",
    
#     # Primary Beneficiary 2 descriptions
#     "ID of second beneficiary (must exist in beneficiary table)",
#     "Allocation percentage for second beneficiary",
#     "Name of second beneficiary",
#     "Identity card/passport number of second beneficiary",
#     "Relationship to client (SPOUSE, CHILD, PARENT, etc.)",
#     "Is beneficiary a minor (TRUE/FALSE)",
#     "Does beneficiary have a guardian (TRUE/FALSE)",
    
#     # Primary Beneficiary 3 descriptions
#     "ID of third beneficiary (must exist in beneficiary table)",
#     "Allocation percentage for third beneficiary",
#     "Name of third beneficiary",
#     "Identity card/passport number of third beneficiary",
#     "Relationship to client (SPOUSE, CHILD, PARENT, etc.)",
#     "Is beneficiary a minor (TRUE/FALSE)",
#     "Does beneficiary have a guardian (TRUE/FALSE)",
    
#     # Guardian Information descriptions
#     "ID of first guardian (must exist in guardian table)",
#     "Name of first guardian",
#     "Identity card/passport number of first guardian",
#     "Relationship to beneficiary (PARENT, LEGAL_GUARDIAN, etc.)",
#     "Beneficiary ID this guardian is assigned to",
#     "ID of second guardian (must exist in guardian table)",
#     "Name of second guardian",
#     "Identity card/passport number of second guardian",
#     "Relationship to beneficiary (PARENT, LEGAL_GUARDIAN, etc.)",
#     "Beneficiary ID this guardian is assigned to"
# ]

# mandatory_product_order_fields = [
#     "order_date", "client_id", "client_type", "product_id", 
#     "order_amount", "unit_price", "payment_method", "payment_status", 
#     "order_status", "sales_channel", "has_beneficiaries"
# ]

# create_headers(ws_product_orders, product_order_headers, product_order_descriptions, mandatory_product_order_fields)

# client_type_dv = DataValidation(type="list", formula1='"INDIVIDUAL,CORPORATE"', allow_blank=True)
# payment_method_dv = DataValidation(type="list", formula1='"BANK_TRANSFER,CREDIT_CARD,CHEQUE,CASH,OTHER"', allow_blank=True)
# payment_status_dv = DataValidation(type="list", formula1='"PENDING,PAID,FAILED,REFUNDED,PARTIAL"', allow_blank=True)
# order_status_dv = DataValidation(type="list", formula1='"PENDING,CONFIRMED,CANCELLED,COMPLETED"', allow_blank=True)
# sales_channel_dv = DataValidation(type="list", formula1='"DIRECT,AGENT,ONLINE,PHONE,EMAIL,WALK_IN,OTHER"', allow_blank=True)
# approval_status_dv = DataValidation(type="list", formula1='"PENDING,APPROVED,REJECTED"', allow_blank=True)
# boolean_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)

# # Apply validations to their respective columns in the Product Orders sheet
# add_sheet_specific_validation(ws_product_orders, "list", '"INDIVIDUAL,CORPORATE"', 'E4:E1000', 'client_type')
# add_sheet_specific_validation(ws_product_orders, "list", '"BANK_TRANSFER,CREDIT_CARD,CHEQUE,CASH,OTHER"', 'J4:J1000', 'payment_method')
# add_sheet_specific_validation(ws_product_orders, "list", '"PENDING,PAID,FAILED,REFUNDED,PARTIAL"', 'K4:K1000', 'payment_status')
# add_sheet_specific_validation(ws_product_orders, "list", '"PENDING,CONFIRMED,CANCELLED,COMPLETED"', 'L4:L1000', 'order_status')
# add_sheet_specific_validation(ws_product_orders, "list", '"DIRECT,AGENT,ONLINE,PHONE,EMAIL,WALK_IN,OTHER"', 'U4:U1000', 'sales_channel')
# add_sheet_specific_validation(ws_product_orders, "list", '"PENDING,APPROVED,REJECTED"', 'X4:X1000', 'approval_status')
# add_sheet_specific_validation(ws_product_orders, "list", '"TRUE,FALSE"', 'Z4:Z1000', 'has_beneficiaries')

# # Add validations for beneficiary and guardian fields
# # Beneficiary 1
# add_sheet_specific_validation(ws_product_orders, "list", '"SELF,FAMILY,ASSOCIATE,CHILD,GUARDIAN,PARENTS,PARTNER,SPOUSE,GRANDPARENT,FRIEND,FIANCE,SIBLING,NIECE,NEPHEW,GRAND_DAUGHTER,GRAND_SON,GOD_PARENT,MOTHER_IN_LAW,FATHER_IN_LAW,SON_IN_LAW,DAUGHTER_IN_LAW"', 'AE4:AE1000', 'beneficiary1_relationship')
# add_sheet_specific_validation(ws_product_orders, "list", '"TRUE,FALSE"', 'AF4:AF1000', 'beneficiary1_is_minor')
# add_sheet_specific_validation(ws_product_orders, "list", '"TRUE,FALSE"', 'AG4:AG1000', 'beneficiary1_has_guardian')

# # Beneficiary 2
# add_sheet_specific_validation(ws_product_orders, "list", '"SELF,FAMILY,ASSOCIATE,CHILD,GUARDIAN,PARENTS,PARTNER,SPOUSE,GRANDPARENT,FRIEND,FIANCE,SIBLING,NIECE,NEPHEW,GRAND_DAUGHTER,GRAND_SON,GOD_PARENT,MOTHER_IN_LAW,FATHER_IN_LAW,SON_IN_LAW,DAUGHTER_IN_LAW"', 'AL4:AL1000', 'beneficiary2_relationship')
# add_sheet_specific_validation(ws_product_orders, "list", '"TRUE,FALSE"', 'AM4:AM1000', 'beneficiary2_is_minor')
# add_sheet_specific_validation(ws_product_orders, "list", '"TRUE,FALSE"', 'AN4:AN1000', 'beneficiary2_has_guardian')

# # Beneficiary 3
# add_sheet_specific_validation(ws_product_orders, "list", '"SELF,FAMILY,ASSOCIATE,CHILD,GUARDIAN,PARENTS,PARTNER,SPOUSE,GRANDPARENT,FRIEND,FIANCE,SIBLING,NIECE,NEPHEW,GRAND_DAUGHTER,GRAND_SON,GOD_PARENT,MOTHER_IN_LAW,FATHER_IN_LAW,SON_IN_LAW,DAUGHTER_IN_LAW"', 'AS4:AS1000', 'beneficiary3_relationship')
# add_sheet_specific_validation(ws_product_orders, "list", '"TRUE,FALSE"', 'AT4:AT1000', 'beneficiary3_is_minor')
# add_sheet_specific_validation(ws_product_orders, "list", '"TRUE,FALSE"', 'AU4:AU1000', 'beneficiary3_has_guardian')

# # Guardian relationships
# add_sheet_specific_validation(ws_product_orders, "list", '"SELF,FAMILY,ASSOCIATE,CHILD,GUARDIAN,PARENTS,PARTNER,SPOUSE,GRANDPARENT,FRIEND,FIANCE,SIBLING,NIECE,NEPHEW,GRAND_DAUGHTER,GRAND_SON,GOD_PARENT,MOTHER_IN_LAW,FATHER_IN_LAW,SON_IN_LAW,DAUGHTER_IN_LAW"', 'AY4:AY1000', 'guardian1_relationship')
# add_sheet_specific_validation(ws_product_orders, "list", '"SELF,FAMILY,ASSOCIATE,CHILD,GUARDIAN,PARENTS,PARTNER,SPOUSE,GRANDPARENT,FRIEND,FIANCE,SIBLING,NIECE,NEPHEW,GRAND_DAUGHTER,GRAND_SON,GOD_PARENT,MOTHER_IN_LAW,FATHER_IN_LAW,SON_IN_LAW,DAUGHTER_IN_LAW"', 'BD4:BD1000', 'guardian2_relationship')

# # ====================== WITHDRAWALS SHEET ======================
# ws_withdrawals = wb.create_sheet(title="Withdrawals")

# withdrawal_headers = [
#     "withdrawal_id", "withdrawal_reference", "client_id", "client_type",
#     "product_id", "order_id", "withdrawal_date", "withdrawal_amount",
#     "withdrawal_units", "unit_price", "withdrawal_status",
#     "bank_account_number", "account_holder_name", "bank_name", "bank_swift_code",
#     "transaction_reference", "transaction_date", "approval_status",
#     "approved_by", "approval_date", "remarks"
# ]

# withdrawal_descriptions = [
#     "Unique identifier for withdrawal (generated if not provided)",
#     "Withdrawal reference number",
#     "ID of client making the withdrawal",
#     "Type of client (INDIVIDUAL, CORPORATE)",
#     "ID of product being withdrawn from",
#     "Original order ID if applicable",
#     "Date withdrawal was requested (YYYY-MM-DD)",
#     "Withdrawal amount in product's currency",
#     "Number of units being withdrawn",
#     "Unit price at time of withdrawal",
#     "Withdrawal status (PENDING, PROCESSING, COMPLETED, REJECTED)",
#     "Bank account number for withdrawal transfer",
#     "Bank account holder name",
#     "Bank name",
#     "Bank SWIFT code",
#     "Bank transaction reference",
#     "Date transaction was completed (YYYY-MM-DD)",
#     "Approval status (PENDING, APPROVED, REJECTED)",
#     "User ID/name who approved/rejected the withdrawal",
#     "Date of approval/rejection (YYYY-MM-DD)",
#     "Additional notes or remarks"
# ]

# mandatory_withdrawal_fields = [
#     "client_id", "client_type", "product_id", "withdrawal_date", 
#     "withdrawal_amount", "withdrawal_units", "unit_price", "withdrawal_status", 
#     "bank_account_number", "account_holder_name", "bank_name", "bank_swift_code", "approval_status"
# ]

# create_headers(ws_withdrawals, withdrawal_headers, withdrawal_descriptions, mandatory_withdrawal_fields)

# withdrawal_status_dv = DataValidation(type="list", formula1='"PENDING,PROCESSING,COMPLETED,REJECTED,CANCELLED"', allow_blank=True)

# # Apply validations to their respective columns in the Withdrawals sheet
# add_sheet_specific_validation(ws_withdrawals, "list", '"INDIVIDUAL,CORPORATE"', 'D4:D1000', 'client_type')
# add_sheet_specific_validation(ws_withdrawals, "list", '"PENDING,PROCESSING,COMPLETED,REJECTED,CANCELLED"', 'K4:K1000', 'withdrawal_status')
# add_sheet_specific_validation(ws_withdrawals, "list", '"PENDING,APPROVED,REJECTED"', 'O4:O1000', 'approval_status')

# # ====================== REDEMPTIONS SHEET ======================
# ws_redemptions = wb.create_sheet(title="Redemptions")

# redemption_headers = [
#     "redemption_id", "redemption_reference", "client_id", "client_type",
#     "product_id", "order_id", "redemption_date", "redemption_amount",
#     "redemption_units", "unit_price", "early_redemption_penalty",
#     "net_amount", "redemption_status", "redemption_reason",
#     "bank_account_number", "account_holder_name", "bank_name", "bank_swift_code",
#     "transaction_reference", "transaction_date", "approval_status",
#     "approved_by", "approval_date", "remarks"
# ]

# redemption_descriptions = [
#     "Unique identifier for redemption (generated if not provided)",
#     "Redemption reference number",
#     "ID of client making the redemption",
#     "Type of client (INDIVIDUAL, CORPORATE)",
#     "ID of product being redeemed",
#     "Original order ID if applicable",
#     "Date redemption was requested (YYYY-MM-DD)",
#     "Redemption amount in product's currency",
#     "Number of units being redeemed",
#     "Unit price at time of redemption",
#     "Early redemption penalty amount if applicable",
#     "Net amount after penalties",
#     "Redemption status (PENDING, PROCESSING, COMPLETED, REJECTED)",
#     "Reason for redemption",
#     "Bank account number for redemption transfer",
#     "Bank account holder name",
#     "Bank name",
#     "Bank SWIFT code",
#     "Bank transaction reference",
#     "Date transaction was completed (YYYY-MM-DD)",
#     "Approval status (PENDING, APPROVED, REJECTED)",
#     "User ID/name who approved/rejected the redemption",
#     "Date of approval/rejection (YYYY-MM-DD)",
#     "Additional notes or remarks"
# ]

# mandatory_redemption_fields = [
#     "client_id", "client_type", "product_id", "redemption_date", 
#     "redemption_amount", "redemption_units", "unit_price", "redemption_status", 
#     "bank_account_number", "account_holder_name", "bank_name", "bank_swift_code", "approval_status"
# ]

# create_headers(ws_redemptions, redemption_headers, redemption_descriptions, mandatory_redemption_fields)

# redemption_status_dv = DataValidation(type="list", formula1='"PENDING,PROCESSING,COMPLETED,REJECTED,CANCELLED"', allow_blank=True)

# # Apply validations to their respective columns in the Redemptions sheet
# add_sheet_specific_validation(ws_redemptions, "list", '"INDIVIDUAL,CORPORATE"', 'D4:D1000', 'client_type')
# add_sheet_specific_validation(ws_redemptions, "list", '"PENDING,PROCESSING,COMPLETED,REJECTED,CANCELLED"', 'M4:M1000', 'redemption_status')
# add_sheet_specific_validation(ws_redemptions, "list", '"PENDING,APPROVED,REJECTED"', 'R4:R1000', 'approval_status')

# # ====================== DIVIDENDS SHEET ======================
# ws_dividends = wb.create_sheet(title="Dividends")

# dividend_headers = [
#     "dividend_id", "dividend_reference", "product_id", "dividend_schedule_id",
#     "payment_date", "record_date", "dividend_rate", "dividend_type",
#     "total_dividend_amount", "status", "remarks"
# ]

# dividend_descriptions = [
#     "Unique identifier for dividend (generated if not provided)",
#     "Dividend reference number",
#     "ID of product this dividend applies to",
#     "ID of the dividend schedule if applicable",
#     "Date dividend payment was/will be made (YYYY-MM-DD)",
#     "Record date for dividend eligibility (YYYY-MM-DD)",
#     "Dividend rate (percentage or fixed amount per unit)",
#     "Dividend type (PERCENTAGE, FIXED_AMOUNT)",
#     "Total dividend amount paid out",
#     "Status (PLANNED, PROCESSING, PAID, CANCELLED)",
#     "Additional notes or remarks"
# ]

# mandatory_dividend_fields = [
#     "product_id", "payment_date", "record_date", "dividend_rate", 
#     "dividend_type", "status"
# ]

# create_headers(ws_dividends, dividend_headers, dividend_descriptions, mandatory_dividend_fields)

# dividend_type_dv = DataValidation(type="list", formula1='"PERCENTAGE,FIXED_AMOUNT"', allow_blank=True)
# dividend_status_dv = DataValidation(type="list", formula1='"PLANNED,PROCESSING,PAID,CANCELLED"', allow_blank=True)

# # Apply validations to their respective columns in the Dividends sheet
# # add_validation_to_sheet(ws_dividends, dividend_type_dv, 'H4:H1000', 'dividend_type')
# add_sheet_specific_validation(ws_dividends, "list", '"PERCENTAGE,FIXED_AMOUNT"', 'H4:H1000', 'dividend_type')
# # add_validation_to_sheet(ws_dividends, dividend_status_dv, 'J4:J1000', 'dividend_status')
# add_sheet_specific_validation(ws_dividends, "list", '"PLANNED,PROCESSING,PAID,CANCELLED"', 'J4:J1000', 'dividend_status')

# # ====================== DIVIDEND PAYMENTS SHEET ======================
# ws_dividend_payments = wb.create_sheet(title="Dividend Payments")

# dividend_payment_headers = [
#     "payment_id", "dividend_id", "client_id", "client_type", "order_id",
#     "units_held", "payment_amount", "payment_date", "payment_status",
#     "bank_account_number", "account_holder_name", "bank_name", "bank_swift_code",
#     "transaction_reference", "transaction_date", "remarks"
# ]

# dividend_payment_descriptions = [
#     "Unique identifier for payment (generated if not provided)",
#     "ID of dividend this payment belongs to",
#     "ID of client receiving payment",
#     "Type of client (INDIVIDUAL, CORPORATE)",
#     "Original order ID",
#     "Number of units held at record date",
#     "Payment amount in product's currency",
#     "Date payment was made (YYYY-MM-DD)",
#     "Payment status (PENDING, PROCESSING, COMPLETED, FAILED)",
#     "Bank account number for dividend payment",
#     "Bank account holder name",
#     "Bank name",
#     "Bank SWIFT code",
#     "Bank transaction reference",
#     "Date transaction was completed (YYYY-MM-DD)",
#     "Additional notes or remarks"
# ]

# mandatory_dividend_payment_fields = [
#     "dividend_id", "client_id", "client_type", "order_id", 
#     "units_held", "payment_amount", "payment_status",
#     "bank_account_number", "account_holder_name", "bank_name", "bank_swift_code"
# ]

# create_headers(ws_dividend_payments, dividend_payment_headers, dividend_payment_descriptions, mandatory_dividend_payment_fields)

# dividend_payment_status_dv = DataValidation(type="list", formula1='"PENDING,PROCESSING,COMPLETED,FAILED,CANCELLED"', allow_blank=True)

# # Apply validations to their respective columns in the Dividend Payments sheet
# add_sheet_specific_validation(ws_dividend_payments, "list", '"INDIVIDUAL,CORPORATE"', 'D4:D1000', 'client_type')
# add_sheet_specific_validation(ws_dividend_payments, "list", '"PLANNED,PROCESSING,PAID,CANCELLED"', 'I4:I1000', 'payment_status')

# # ====================== COMMISSIONS SHEET ======================
# ws_commissions = wb.create_sheet(title="Commissions")

# commission_headers = [
#     "commission_id", "commission_reference", "agent_id", "agency_id",
#     "commission_date", "commission_type", "product_id", "order_id", "client_id",
#     "base_amount", "rate_percentage", "commission_amount", "status",
#     "payment_date", "payment_reference", "bank_account_number", "bank_name", "bank_swift_code",
#     "remarks"
# ]

# commission_descriptions = [
#     "Unique identifier for commission (generated if not provided)",
#     "Commission reference number",
#     "ID of agent receiving the commission",
#     "ID of agency the agent belongs to",
#     "Date commission was calculated (YYYY-MM-DD)",
#     "Type of commission (SALES, RECRUITMENT, OVERRIDE, BONUS, etc.)",
#     "ID of product related to this commission",
#     "ID of order related to this commission if applicable",
#     "ID of client related to this commission if applicable",
#     "Base amount for commission calculation",
#     "Commission rate percentage",
#     "Commission amount in local currency",
#     "Status (PENDING, APPROVED, PAID, REJECTED, CANCELLED)",
#     "Date payment was made (YYYY-MM-DD)",
#     "Payment reference number",
#     "Bank account number for commission payment",
#     "Bank name",
#     "Bank SWIFT code",
#     "Additional notes or remarks"
# ]

# mandatory_commission_fields = [
#     "agent_id", "agency_id", "commission_date", "commission_type", 
#     "base_amount", "rate_percentage", "commission_amount", "status",
#     "bank_account_number", "bank_name", "bank_swift_code"
# ]

# create_headers(ws_commissions, commission_headers, commission_descriptions, mandatory_commission_fields)

# commission_status_dv = DataValidation(type="list", formula1='"PENDING,APPROVED,PAID,REJECTED,CANCELLED"', allow_blank=True)

# # Apply validations to their respective columns in the Commissions sheet
# add_sheet_specific_validation(ws_commissions, "list", '"SALES,RECRUITMENT,OVERRIDE,BONUS,MANAGEMENT,RENEWAL"', 'F4:F1000', 'commission_type')
# add_sheet_specific_validation(ws_commissions, "list", '"PENDING,APPROVED,PAID,REJECTED,CANCELLED"', 'M4:M1000', 'commission_status')

# Save the workbook
filename = "citadel_data_migration_template.xlsx"
wb.save(filename)

print(f"Excel template created: {filename}")

# Display a summary of the sheets included
print("\nTemplate includes the following sheets:")
for sheet in wb.sheetnames:
    print(f"- {sheet}")